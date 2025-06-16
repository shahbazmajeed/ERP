from openpyxl import load_workbook
from django.contrib import messages
from django.shortcuts import render, redirect
from ..models import CustomUser, Employee, Department
from ..forms import UploadEmployeeFileForm
from django.contrib.auth.decorators import login_required

from django.contrib.auth import get_user_model

import openpyxl

User = get_user_model()

@login_required
def upload_employees(request):
    if request.user.role != 'admin':
        return render(request, 'error.html', {'message': 'Unauthorized access'})

    if request.method == 'POST':
        form = UploadEmployeeFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            added_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 10:
                    continue  # skip incomplete rows

                eid, username, first_name, last_name, email, password, role, dept_name, designation, position_type = row

                # Skip existing usernames or employee IDs
                if User.objects.filter(username=username).exists() or Employee.objects.filter(eid=eid).exists():
                    continue

                # Get or create the department
                department, _ = Department.objects.get_or_create(name=dept_name)

                # Create user
                user = User.objects.create_user(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    password=password,
                    role=role
                )

                # Create employee profile
                Employee.objects.create(
                    user=user,
                    eid=eid,
                    department=department,
                    designation=designation,
                    position_type=position_type
                )

                added_count += 1

            messages.success(request, f"Successfully uploaded {added_count} employees.")
            return redirect('list_employees')

    else:
        form = UploadEmployeeFileForm()

    return render(request, 'upload_employees.html', {'form': form})

@login_required
def list_employees(request):
    if request.user.role not in ['admin', 'director', 'hod']:
        return render(request, 'error.html', {'message': 'Unauthorized access'})

    employees = Employee.objects.select_related('user', 'department')

    # Optional filters
    filter_dept = request.GET.get('department', '')
    filter_position = request.GET.get('position', '')

    if filter_dept:
        employees = employees.filter(department__name=filter_dept)
    if filter_position:
        employees = employees.filter(position_type=filter_position)

    departments = Department.objects.values_list('name', flat=True).distinct()
    positions = Employee.objects.values_list('position_type', flat=True).distinct()

    return render(request, 'list_employees.html', {
        'employees': employees,
        'departments': departments,
        'positions': positions,
        'filter_dept': filter_dept,
        'filter_position': filter_position,
    })
